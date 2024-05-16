import openpyxl


def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_col_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(path, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value


def set_cell_data(path, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(path)


path = "..//excel//testdata.xlsx"
shet_name = "LoginTest"

rows = get_row_count(path, shet_name)
columns = get_col_count(path, shet_name)

print(rows, "---", columns)

print(get_cell_data(path, shet_name, 2, 1))
set_cell_data(path, shet_name, 1, 4, "DOB")
