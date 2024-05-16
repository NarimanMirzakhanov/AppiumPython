import openpyxl


def get_data(sheet_name):

    workbook = openpyxl.load_workbook("/Users/Nariman/PycharmProjects/AppiumPageObjectModel/excel/testdata2.xlsx")
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    print("total rows are ", str(total_rows))
    print("total columns are ", str(total_columns))
    main_list = []

    for i in range(2, total_rows + 1):
        data_list = []
        for j in range(1, total_columns + 1):
            data = sheet.cell(row=i, column=j).value
            data_list.insert(j, data)
        main_list.insert(i, data_list)
    return main_list
